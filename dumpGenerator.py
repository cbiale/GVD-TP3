#! python3
import openpyxl
wb = openpyxl.load_workbook('rawData.xlsx')
wb = openpyxl.load_workbook('rawData.xlsx')

subjects = []
persons = []
relations = []

wsnumber = 0
pIndex = 0
for ws in wb:
    code = ws.title
    rawSubject = ws['B' + str(len(ws['B']))].value
    [name, schedule] = rawSubject.split('-')
    email = ws['D' + str(len(ws['D']))].value
    id = 's%d' % (wsnumber)
    subjects.append({'code': code, 'name': name,
                     'email': email, 'schedule': schedule})
    for row in ws.iter_rows(min_row=2):
        if(row[0].value != 'Materia'):
            pFirstName = row[1].value
            pLastName = row[2].value
            pFullName = '%s, %s' % (pLastName, pFirstName)
            pEmail = row[3].value
            pId = 'p%d' % (pIndex)
            persons.append({'name': pFullName, 'email': pEmail})
            relations.append({'pEmail': pEmail, 'sEmail': email,
                              'relationship': row[0].value, 'grade': row[4].value})
            pIndex += 1
    wsnumber += 1

uniquePersons = [dict(t) for t in {tuple(d.items()) for d in persons}]

dump = open("dump.cypher", "w")

dump.write('CREATE \n')

dump.write('(g0:Group {name:"alfa"}),\n')
dump.write('(g1:Group {name:"bravo"}),\n')
dump.write('(g2:Group {name:"charlie"}),\n')

i = 0
for s in subjects:
    dump.write('(s%d:Subject {code: "%s",name:"%s", email:"%s", schedule:"%s"}),\n' % (
        i, s['code'], s['name'], s['email'], s['schedule']))
    i += 1

i = 0
for p in uniquePersons:
    dump.write('(p%d:Person {name: "%s", email:"%s"}),\n' %
               (i, p['name'], p['email']))
    i += 1

pno = 0
i = 0
for p in uniquePersons:
    dump.write('(p%d)-[:BELONGS]->(g%d),\n' % (pno, i))
    pno += 1
    if (i == 2):
        i = 0
    else:
        i += 1

for r in relations:
    if(r['relationship'] == 'Docente'):
        dump.write(
            'MATCH (auxP: Person {email:"%s"}), (auxS: Subject {email:"%s"}) CREATE (auxP)-[:TEACH]->(auxS);\n' % (r['pEmail'], r['sEmail']))
    if(r['relationship'] == 'Alumno'):
        dump.write(
            'MATCH (auxP: Person {email:"%s"}), (auxS: Subject {email:"%s"}) CREATE (auxP)-[:STUDIED {grade:"%d"}]->(auxS);\n' % (r['pEmail'], r['sEmail'],r['grade']))
    
dump.close()
